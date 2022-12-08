import pandas as pd
import openpyxl
from openpyxl.writer.excel import save_workbook
import random


# Получение фрейма
# В пути должен быть начальный эксель файл
started_file_path = 'assets/started_file.xlsx'
data_frame = pd.read_excel(started_file_path)

# Формирование колонок и значений

columns = list(data_frame.columns.values)

if columns[0] / columns[0] == 1.0:
    for i in columns:
        summand = random.uniform(-i * 0.1, i * 0.1)
        columns[columns.index(i)] = round(i + summand, 3)
print(columns)

rows = []

for i in data_frame.values.tolist():
    for j in i:
        summand = random.uniform(-j * 0.1, j * 0.1)
        i[i.index(j)] = round(j + summand, 3)
    if i[0] / i[0] == 1.0:
        print(i)
        rows.append(i)

# Создание файла
finished_file_path = 'assets/finished_file.xlsx'

try:
    wb = openpyxl.load_workbook(finished_file_path)

except:
    wb = openpyxl.Workbook()

    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet)

ws = wb.create_sheet('1')

for i, value in enumerate(columns, 1):
    ws.cell(row=1, column=i).value = value

for i, row in enumerate(rows, 2):
    for j, value in enumerate(row, 1):
        ws.cell(row=i, column=j).value = value

save_workbook(wb, finished_file_path)